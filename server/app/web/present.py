import csv
import io
import json
import pathlib
from typing import Any

from fastapi import APIRouter, Depends, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .. import models
from ..db import get_db
from ..errors import ApiError
from ..eval.metrics import compute_metrics
from ..eval.service import _parse_sessions
from .detail import _extract_dir_for_attempt, _final_grade, _latest_evaluation
from .pages import get_teacher_page

router = APIRouter()

_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}


def _override_for_student(db: Session, assignment_id: int, student_id: int):
    return (
        db.query(models.GradeOverride)
        .filter_by(
            target_type="individual",
            target_id=f"{assignment_id}:{student_id}",
        )
        .order_by(models.GradeOverride.updated_at.desc())
        .first()
    )


def _group_override(db: Session, assignment_id: int, group_id: int):
    return (
        db.query(models.GradeOverride)
        .filter_by(
            target_type="group",
            target_id=f"{assignment_id}:{group_id}",
        )
        .order_by(models.GradeOverride.updated_at.desc())
        .first()
    )


def _group_screenshots(
    attempts: list[models.SubmissionAttempt], settings: Any
) -> list[str]:
    urls: list[str] = []
    for attempt in attempts:
        extract_dir = pathlib.Path(_extract_dir_for_attempt(attempt, settings))
        screenshots_dir = extract_dir / "screenshots"
        if not screenshots_dir.exists():
            continue
        for p in sorted(screenshots_dir.iterdir()):
            if p.is_file() and p.suffix.lower() in _IMAGE_SUFFIXES:
                rel = p.relative_to(extract_dir).as_posix()
                urls.append(f"/media/extracted/{attempt.id}/{rel}")
    return urls


def _group_code_highlights(
    attempts: list[models.SubmissionAttempt], settings: Any
) -> list[str]:
    """Return a compact, non-executable project file summary for projection."""
    names: list[str] = []
    seen: set[str] = set()
    for attempt in attempts:
        code_dir = pathlib.Path(_extract_dir_for_attempt(attempt, settings)) / "code"
        if not code_dir.exists():
            continue
        for path in sorted(code_dir.rglob("*")):
            if path.is_file():
                rel = path.relative_to(code_dir).as_posix()
                if rel not in seen:
                    seen.add(rel)
                    names.append(rel)
                if len(names) >= 10:
                    return names
    return names


def _group_metrics(
    attempts: list[models.SubmissionAttempt], settings: Any
) -> dict[str, Any]:
    all_timelines: list[Any] = []
    for attempt in attempts:
        extract_dir = _extract_dir_for_attempt(attempt, settings)
        all_timelines.extend(_parse_sessions(extract_dir))
    metrics = compute_metrics(all_timelines)
    return {
        "sessions": metrics.get("sessions", 0),
        "turns": metrics.get("turns", 0),
        "duration_min": metrics.get("duration_min", 0),
    }


def _build_group_row(
    db: Session,
    assignment: models.Assignment,
    group: models.Group | None,
    students: list[models.Student],
    settings: Any,
) -> dict[str, Any]:
    members: list[dict[str, Any]] = []
    attempts: list[models.SubmissionAttempt] = []
    for student in students:
        submission = (
            db.query(models.Submission)
            .filter_by(assignment_id=assignment.id, student_id=student.id)
            .first()
        )
        attempt = None
        if submission is not None and submission.current_attempt_id is not None:
            attempt = db.get(models.SubmissionAttempt, submission.current_attempt_id)
        members.append(
            {
                "name": student.name,
                "student_no": student.student_no,
            }
        )
        if attempt is not None:
            attempts.append(attempt)

    return {
        "group_name": group.name if group else "未分组",
        "members": members,
        "screenshots": _group_screenshots(attempts, settings),
        "metrics": _group_metrics(attempts, settings),
        "code_highlights": _group_code_highlights(attempts, settings),
    }


def present_data(db: Session, assignment: models.Assignment, settings: Any) -> list[dict[str, Any]]:
    groups = (
        db.query(models.Group)
        .filter_by(course_id=assignment.course_id)
        .order_by(models.Group.name)
        .all()
    )

    rows: list[dict[str, Any]] = []
    for group in groups:
        students = (
            db.query(models.Student)
            .filter_by(group_id=group.id, course_id=assignment.course_id)
            .order_by(models.Student.student_no)
            .all()
        )
        if students:
            rows.append(_build_group_row(db, assignment, group, students, settings))

    ungrouped = (
        db.query(models.Student)
        .filter_by(course_id=assignment.course_id, group_id=None)
        .order_by(models.Student.student_no)
        .all()
    )
    if ungrouped:
        rows.append(
            _build_group_row(db, assignment, None, ungrouped, settings)
        )

    return rows


@router.get("/assignments/{aid}/present", response_class=HTMLResponse)
def present_page(
    request: Request,
    aid: int,
    db: Session = Depends(get_db),
    t: models.Teacher = Depends(get_teacher_page),
):
    from ..main import templates

    assignment = db.get(models.Assignment, aid)
    if assignment is None:
        raise ApiError(404, "NOT_FOUND", "作业不存在")

    settings = request.app.state.settings
    groups = present_data(db, assignment, settings)

    return templates.TemplateResponse(
        request,
        "present.html",
        {
            "teacher": t,
            "assignment": assignment,
            "groups": groups,
        },
    )


def review_present_data(
    db: Session, assignment: models.Assignment
) -> list[dict[str, Any]]:
    """Teacher-controlled evaluation showcase, intentionally separate from work projection."""
    groups = (
        db.query(models.Group)
        .filter_by(course_id=assignment.course_id)
        .order_by(models.Group.name)
        .all()
    )
    rows: list[dict[str, Any]] = []
    group_specs: list[tuple[models.Group | None, list[models.Student]]] = []
    for group in groups:
        students = (
            db.query(models.Student)
            .filter_by(course_id=assignment.course_id, group_id=group.id)
            .order_by(models.Student.student_no)
            .all()
        )
        if students:
            group_specs.append((group, students))
    ungrouped = (
        db.query(models.Student)
        .filter_by(course_id=assignment.course_id, group_id=None)
        .order_by(models.Student.student_no)
        .all()
    )
    if ungrouped:
        group_specs.append((None, ungrouped))

    for group, students in group_specs:
        members = []
        for student in students:
            submission = (
                db.query(models.Submission)
                .filter_by(assignment_id=assignment.id, student_id=student.id)
                .first()
            )
            attempt = db.get(models.SubmissionAttempt, submission.current_attempt_id) if submission and submission.current_attempt_id else None
            evaluation = _latest_evaluation(db, attempt)
            override = _override_for_student(db, assignment.id, student.id)
            members.append(
                {
                    "name": student.name,
                    "student_no": student.student_no,
                    "status": submission.status if submission else "未提交",
                    "final_grade": _final_grade(override, evaluation),
                    "rationale": evaluation.rationale if evaluation else "",
                    "feedback": evaluation.feedback_json if evaluation else [],
                    "teacher_note": override.comment if override and not override.stale else "",
                }
            )
        geval = None
        group_override = None
        if group is not None:
            geval = (
                db.query(models.GroupEvaluation)
                .filter_by(assignment_id=assignment.id, group_id=group.id)
                .order_by(models.GroupEvaluation.created_at.desc())
                .first()
            )
            group_override = _group_override(db, assignment.id, group.id)
        rows.append(
            {
                "group_name": group.name if group else "未分组",
                "members": members,
                "group_final_grade": _final_grade(group_override, geval),
                "group_rationale": geval.rationale if geval else "",
            }
        )
    return rows


@router.get("/assignments/{aid}/review-present", response_class=HTMLResponse)
def review_present_page(
    request: Request,
    aid: int,
    db: Session = Depends(get_db),
    t: models.Teacher = Depends(get_teacher_page),
):
    from ..main import templates

    assignment = db.get(models.Assignment, aid)
    if assignment is None:
        raise ApiError(404, "NOT_FOUND", "作业不存在")
    return templates.TemplateResponse(
        request,
        "review_present.html",
        {"teacher": t, "assignment": assignment, "groups": review_present_data(db, assignment)},
    )


def _csv_row(
    db: Session,
    assignment_id: int,
    student: models.Student,
    group_name: str,
    group_evaluation: models.GroupEvaluation | None = None,
    group_override: models.GradeOverride | None = None,
) -> list[str]:
    submission = (
        db.query(models.Submission)
        .filter_by(assignment_id=assignment_id, student_id=student.id)
        .first()
    )
    attempt = None
    if submission is not None and submission.current_attempt_id is not None:
        attempt = db.get(models.SubmissionAttempt, submission.current_attempt_id)
    evaluation = _latest_evaluation(db, attempt)
    override = _override_for_student(db, assignment_id, student.id)

    ai_grade = evaluation.grade if evaluation else ""
    final_grade = _final_grade(override, evaluation) or ""

    dims: dict[str, Any] = {}
    if evaluation is not None and evaluation.dimension_scores_json:
        for d in evaluation.dimension_scores_json:
            if isinstance(d, dict) and "name" in d and "score" in d:
                dims[d["name"]] = d["score"]
    dim_json = json.dumps(dims, ensure_ascii=False, separators=(",", ":")) if dims else ""

    status = submission.status if submission else "none"
    feedback = "\n".join(f"• {item}" for item in (evaluation.feedback_json if evaluation else []))
    group_grade = _final_grade(group_override, group_evaluation) or ""
    group_rationale = group_evaluation.rationale if group_evaluation else ""
    teacher_note = override.comment if override and not override.stale else ""
    return [
        student.student_no,
        student.name,
        group_name,
        status,
        ai_grade,
        final_grade,
        evaluation.rationale if evaluation else "",
        feedback,
        teacher_note,
        group_grade,
        group_rationale,
        dim_json,
    ]


def _safe_csv_cell(value: Any) -> str:
    """Prevent spreadsheet formula execution when feedback is opened in Excel."""
    text = "" if value is None else str(value)
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _assignment_export_records(
    db: Session, assignment: models.Assignment
) -> list[dict[str, Any]]:
    """Return structured export records once, for both CSV and formatted Excel."""
    records: list[dict[str, Any]] = []
    groups = (
        db.query(models.Group)
        .filter_by(course_id=assignment.course_id)
        .order_by(models.Group.name)
        .all()
    )
    group_specs: list[tuple[models.Group | None, list[models.Student]]] = []
    for group in groups:
        students = (
            db.query(models.Student)
            .filter_by(group_id=group.id, course_id=assignment.course_id)
            .order_by(models.Student.student_no)
            .all()
        )
        if students:
            group_specs.append((group, students))
    ungrouped = (
        db.query(models.Student)
        .filter_by(course_id=assignment.course_id, group_id=None)
        .order_by(models.Student.student_no)
        .all()
    )
    if ungrouped:
        group_specs.append((None, ungrouped))

    for group, students in group_specs:
        group_evaluation = None
        group_override = None
        if group is not None:
            group_evaluation = (
                db.query(models.GroupEvaluation)
                .filter_by(assignment_id=assignment.id, group_id=group.id)
                .order_by(models.GroupEvaluation.created_at.desc())
                .first()
            )
            group_override = _group_override(db, assignment.id, group.id)
        for student in students:
            submission = (
                db.query(models.Submission)
                .filter_by(assignment_id=assignment.id, student_id=student.id)
                .first()
            )
            attempt = (
                db.get(models.SubmissionAttempt, submission.current_attempt_id)
                if submission and submission.current_attempt_id
                else None
            )
            evaluation = _latest_evaluation(db, attempt)
            override = _override_for_student(db, assignment.id, student.id)
            records.append(
                {
                    "student": student,
                    "group_name": group.name if group else "未分组",
                    "submission": submission,
                    "evaluation": evaluation,
                    "override": override,
                    "group_evaluation": group_evaluation,
                    "group_override": group_override,
                }
            )
    return records


def _safe_excel_cell(value: Any) -> str:
    """Excel applies formulas to leading operators, so retain the CSV protection."""
    return _safe_csv_cell(value)


def _style_export_sheet(sheet, assignment: models.Assignment, headers: list[str], widths: list[int]) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(1, 1, f"Vibe 作业反馈表 · {assignment.title}")
    title.font = Font(bold=True, color="FFFFFF", size=16)
    title.fill = PatternFill("solid", fgColor="0F172A")
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sheet.cell(2, 1, f"作业码：{assignment.code}    截止时间：{assignment.deadline}")
    sheet.cell(2, 1).font = Font(color="075985", italic=True, bold=True)
    sheet.cell(2, 1).fill = PatternFill("solid", fgColor="E0F2FE")
    sheet.cell(2, 1).alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 22
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(4, index, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F766E")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = widths[index - 1]
    sheet.row_dimensions[4].height = 26
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    sheet.sheet_view.showGridLines = False


def _write_export_row(sheet, row_index: int, values: list[Any]) -> None:
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column, _safe_excel_cell(value))
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
    if row_index % 2:
        for cell in sheet[row_index]:
            cell.fill = PatternFill("solid", fgColor="F8FAFC")


def _dimension_summary(evaluation: models.Evaluation | None) -> str:
    lines: list[str] = []
    for dimension in (evaluation.dimension_scores_json if evaluation else []):
        if not isinstance(dimension, dict):
            continue
        name = dimension.get("name", "未命名维度")
        score = dimension.get("score", "")
        weight = dimension.get("weight", "")
        rationale = dimension.get("rationale", "")
        lines.append(f"• {name}｜{score} 分｜权重 {weight}%\n  {rationale}".rstrip())
    return "\n\n".join(lines)


@router.get("/assignments/{aid}/export.xlsx")
def export_xlsx(
    request: Request,
    aid: int,
    db: Session = Depends(get_db),
    t: models.Teacher = Depends(get_teacher_page),
):
    """Export a teacher-readable workbook instead of forcing long feedback into CSV."""
    assignment = db.get(models.Assignment, aid)
    if assignment is None:
        raise ApiError(404, "NOT_FOUND", "作业不存在")
    records = _assignment_export_records(db, assignment)
    workbook = Workbook()
    feedback = workbook.active
    feedback.title = "个人反馈"
    feedback_headers = [
        "学号", "姓名", "小组", "提交状态", "AI 等级", "最终等级",
        "个人评价", "改进建议", "评分维度与说明", "教师备注",
    ]
    _style_export_sheet(
        feedback,
        assignment,
        feedback_headers,
        [14, 12, 12, 13, 10, 10, 42, 42, 54, 26],
    )
    for row_index, record in enumerate(records, start=5):
        submission = record["submission"]
        evaluation = record["evaluation"]
        override = record["override"]
        _write_export_row(
            feedback,
            row_index,
            [
                record["student"].student_no,
                record["student"].name,
                record["group_name"],
                submission.status if submission else "未提交",
                evaluation.grade if evaluation else "",
                _final_grade(override, evaluation) or "",
                evaluation.rationale if evaluation else "",
                "\n".join(f"• {item}" for item in (evaluation.feedback_json if evaluation else [])),
                _dimension_summary(evaluation),
                override.comment if override and not override.stale else "",
            ],
        )
        feedback.row_dimensions[row_index].height = 150

    group_sheet = workbook.create_sheet("小组反馈")
    group_headers = ["小组", "成员", "最终等级", "小组评价"]
    _style_export_sheet(group_sheet, assignment, group_headers, [20, 36, 12, 82])
    group_rows: dict[str, dict[str, Any]] = {}
    for record in records:
        group_rows.setdefault(record["group_name"], record)
    for row_index, record in enumerate(group_rows.values(), start=5):
        group_evaluation = record["group_evaluation"]
        members = "、".join(
            f"{item['student'].name}（{item['student'].student_no}）"
            for item in records
            if item["group_name"] == record["group_name"]
        )
        _write_export_row(
            group_sheet,
            row_index,
            [
                record["group_name"],
                members,
                _final_grade(record["group_override"], group_evaluation) or "",
                group_evaluation.rationale if group_evaluation else "",
            ],
        )
        group_sheet.row_dimensions[row_index].height = 96

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="vibe-feedback.xlsx"'},
    )


@router.get("/assignments/{aid}/export.csv")
def export_csv(
    request: Request,
    aid: int,
    db: Session = Depends(get_db),
    t: models.Teacher = Depends(get_teacher_page),
):
    assignment = db.get(models.Assignment, aid)
    if assignment is None:
        raise ApiError(404, "NOT_FOUND", "作业不存在")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "学号", "姓名", "小组", "提交状态", "AI等级", "最终等级",
        "个人评价", "个人改进建议", "教师备注", "小组最终等级", "小组评价", "各维度分(json)",
    ])

    groups = (
        db.query(models.Group)
        .filter_by(course_id=assignment.course_id)
        .order_by(models.Group.name)
        .all()
    )
    for group in groups:
        group_evaluation = (
            db.query(models.GroupEvaluation)
            .filter_by(assignment_id=assignment.id, group_id=group.id)
            .order_by(models.GroupEvaluation.created_at.desc())
            .first()
        )
        group_override = _group_override(db, assignment.id, group.id)
        students = (
            db.query(models.Student)
            .filter_by(group_id=group.id, course_id=assignment.course_id)
            .order_by(models.Student.student_no)
            .all()
        )
        for student in students:
            writer.writerow([
                _safe_csv_cell(cell)
                for cell in _csv_row(
                    db, assignment.id, student, group.name, group_evaluation, group_override
                )
            ])

    ungrouped = (
        db.query(models.Student)
        .filter_by(course_id=assignment.course_id, group_id=None)
        .order_by(models.Student.student_no)
        .all()
    )
    for student in ungrouped:
        writer.writerow([_safe_csv_cell(cell) for cell in _csv_row(db, assignment.id, student, "未分组")])

    content = output.getvalue()
    encoded = content.encode("utf-8")
    return StreamingResponse(
        io.BytesIO(encoded),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="export.csv"'},
    )
