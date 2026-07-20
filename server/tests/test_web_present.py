"""TDD tests for Task 3: presentation mode and CSV export."""
import csv
import hashlib
import io
import json
import re
import zipfile
from datetime import timedelta

from openpyxl import load_workbook

from app import models
from app.db import SessionLocal
from app.eval.worker import run_worker_once
from app.utils import utcnow
from tests.test_auth import _mk_teacher
from tests.test_eval_pipeline import (
    FakeLLMProvider,
    _valid_group_json,
    _valid_individual_json,
)


def _login(client):
    _mk_teacher()
    return client.post("/login", json={"username": "admin", "password": "pw123456"})


def _setup_course_and_assignment(client):
    _login(client)
    cid = client.post("/courses", json={"name": "C", "term": ""}).json()["id"]
    csv_text = "学号,姓名,小组\n2024001,张三,G1\n2024002,李四,G1\n2024003,王五,G2\n"
    body = client.post(f"/courses/{cid}/roster", json={"csv": csv_text}).json()
    tokens = {}
    for line in body["tokens_csv"].splitlines()[1:]:
        no, _name, token = line.split(",")
        tokens[no] = token
    now = utcnow()
    code = client.post(
        f"/courses/{cid}/assignments",
        json={
            "title": "A",
            "description": "",
            "rubric": [
                {"name": "需求理解", "weight": 30, "description": ""},
                {"name": "实现质量", "weight": 40, "description": ""},
                {"name": "迭代能力", "weight": 30, "description": ""},
            ],
            "opens_at": (now - timedelta(days=1)).isoformat(),
            "deadline": (now + timedelta(days=1)).isoformat(),
            "max_package_mb": 50,
        },
    ).json()["code"]
    return cid, tokens, code


def _package(code, student_no):
    files = {
        "sessions/sess-abc.jsonl": json.dumps(
            {
                "type": "response_item",
                "payload": {
                    "type": "message",
                    "role": "user",
                    "content": [{"type": "input_text", "text": "Implement a todo list"}],
                    "timestamp": "2026-07-19T10:01:00+08:00",
                },
            },
            ensure_ascii=False,
        ).encode()
        + b"\n"
        + json.dumps(
            {
                "type": "response_item",
                "payload": {
                    "type": "message",
                    "role": "assistant",
                    "content": [{"type": "output_text", "text": "I'll create a FastAPI app."}],
                    "timestamp": "2026-07-19T10:02:00+08:00",
                },
            },
            ensure_ascii=False,
        ).encode()
        + b"\n",
        "code/main.py": b"print(1)",
        "screenshots/sc.png": b"fake-image-data",
    }
    manifest = {
        "format_version": "1",
        "assignment_code": code,
        "student_no": student_no,
        "client_version": "0.1.0",
        "submitted_at": "2026-07-19T08:00:00Z",
        "files": [{"path": n, "sha256": hashlib.sha256(b).hexdigest()} for n, b in files.items()],
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
        for n, b in files.items():
            z.writestr(n, b)
    return manifest, buf.getvalue()


def _upload(client, token, code, student_no):
    manifest, blob = _package(code, student_no)
    data = {"manifest": json.dumps(manifest, ensure_ascii=False)}
    return client.post(
        "/api/submissions",
        headers={"Authorization": f"Bearer {token}"},
        data=data,
        files={"file": ("p.zip", blob, "application/zip")},
    )


class TestPresentPage:
    def test_present_redirects_when_unauthenticated(self, client):
        r = client.get("/assignments/1/present", follow_redirects=False)
        assert r.status_code == 302
        assert "/login" in r.headers["location"]

    def test_present_404_for_missing_assignment(self, client):
        _login(client)
        r = client.get("/assignments/999/present")
        assert r.status_code == 404

    def test_work_present_shows_student_work_without_evaluation(self, client, settings):
        _setup_course_and_assignment(client)
        db = SessionLocal()
        assignment = db.query(models.Assignment).one()
        students = db.query(models.Student).order_by(models.Student.student_no).all()
        s1, s2 = students[0], students[1]
        from app.security import hash_token, new_submit_token
        t1 = new_submit_token()
        t2 = new_submit_token()
        s1.submit_token_hash = hash_token(t1)
        s2.submit_token_hash = hash_token(t2)
        db.commit()
        db.close()

        assert _upload(client, t1, assignment.code, s1.student_no).status_code == 201
        assert _upload(client, t2, assignment.code, s2.student_no).status_code == 201

        provider = FakeLLMProvider(
            responses=[
                _valid_individual_json(),
                _valid_individual_json(),
                _valid_group_json(),
            ]
        )
        db = SessionLocal()
        run_worker_once(db, provider, settings)
        db.close()
        db = SessionLocal()
        run_worker_once(db, provider, settings)
        db.close()

        db = SessionLocal()
        aid = db.query(models.Assignment).one().id
        db.close()

        r = client.get(f"/assignments/{aid}/present")
        assert r.status_code == 200
        text = r.text
        assert "G1" in text
        assert "作品展示" in text
        assert "main.py" in text
        assert "小组 AI 等级" not in text
        assert "小组整体表现良好" not in text
        match = re.search(r"const GROUPS\s*=\s*(.*?);", text, re.DOTALL)
        embedded = json.loads(match.group(1))
        g1 = next(item for item in embedded if item["group_name"] == "G1")
        assert {member["name"] for member in g1["members"]} == {"张三", "李四"}

    def test_present_embedded_json_excludes_feedback_flags_and_dimensions(self, client, settings):
        _setup_course_and_assignment(client)
        db = SessionLocal()
        assignment = db.query(models.Assignment).one()
        students = db.query(models.Student).order_by(models.Student.student_no).all()
        s1, s2 = students[0], students[1]
        from app.security import hash_token, new_submit_token
        t1 = new_submit_token()
        t2 = new_submit_token()
        s1.submit_token_hash = hash_token(t1)
        s2.submit_token_hash = hash_token(t2)
        db.commit()
        db.close()

        assert _upload(client, t1, assignment.code, s1.student_no).status_code == 201
        assert _upload(client, t2, assignment.code, s2.student_no).status_code == 201

        provider = FakeLLMProvider(
            responses=[
                _valid_individual_json(),
                _valid_individual_json(),
                _valid_group_json(),
            ]
        )
        db = SessionLocal()
        run_worker_once(db, provider, settings)
        db.close()
        db = SessionLocal()
        run_worker_once(db, provider, settings)
        db.close()

        db = SessionLocal()
        aid = db.query(models.Assignment).one().id
        db.close()

        r = client.get(f"/assignments/{aid}/present")
        text = r.text

        match = re.search(r"const GROUPS\s*=\s*(.*?);", text, re.DOTALL)
        assert match, "GROUPS JSON not embedded"
        embedded = json.loads(match.group(1))
        assert isinstance(embedded, list)
        g1 = next(item for item in embedded if item["group_name"] == "G1")
        assert "highlight" not in g1
        assert "final_grade" not in json.dumps(g1)
        assert "feedback" not in json.dumps(g1)
        assert "flags" not in json.dumps(g1)
        assert "dimension_scores" not in json.dumps(g1)

        # The raw page should not leak forbidden content values.
        assert "注意代码风格统一" not in text
        assert "无真实性风险" not in text

    def test_review_presentation_has_evaluation_and_feedback(self, client, settings):
        _setup_course_and_assignment(client)
        db = SessionLocal()
        assignment = db.query(models.Assignment).one()
        student = db.query(models.Student).order_by(models.Student.student_no).first()
        from app.security import hash_token, new_submit_token
        token = new_submit_token()
        student.submit_token_hash = hash_token(token)
        db.commit()
        db.close()

        assert _upload(client, token, assignment.code, student.student_no).status_code == 201
        db = SessionLocal()
        run_worker_once(db, FakeLLMProvider(responses=[_valid_individual_json()]), settings)
        db.close()

        r = client.get(f"/assignments/{assignment.id}/review-present")
        assert r.status_code == 200
        assert "评价展示" in r.text
        assert "改进建议" in r.text
        assert "更多功能测试" in r.text


class TestExportCsv:
    def test_export_requires_teacher(self, client):
        r = client.get("/assignments/1/export.csv", follow_redirects=False)
        assert r.status_code == 302
        assert "/login" in r.headers["location"]

    def test_export_404_for_missing_assignment(self, client):
        _login(client)
        r = client.get("/assignments/999/export.csv")
        assert r.status_code == 404

    def test_export_csv_has_header_and_grade_rows(self, client, settings):
        _setup_course_and_assignment(client)
        db = SessionLocal()
        assignment = db.query(models.Assignment).one()
        s1 = db.query(models.Student).order_by(models.Student.student_no).first()
        from app.security import hash_token, new_submit_token
        token = new_submit_token()
        s1.submit_token_hash = hash_token(token)
        db.commit()
        db.close()

        assert _upload(client, token, assignment.code, s1.student_no).status_code == 201

        provider = FakeLLMProvider(responses=[_valid_individual_json()])
        db = SessionLocal()
        run_worker_once(db, provider, settings)
        db.close()

        db = SessionLocal()
        aid = db.query(models.Assignment).one().id
        db.close()

        r = client.get(f"/assignments/{aid}/export.csv")
        assert r.status_code == 200
        assert "text/csv" in r.headers["content-type"]
        assert 'attachment' in r.headers["content-disposition"]

        reader = csv.reader(io.StringIO(r.text))
        rows = list(reader)
        assert rows[0] == [
            "学号", "姓名", "小组", "提交状态", "AI等级", "最终等级",
            "个人评价", "个人改进建议", "教师备注", "小组最终等级", "小组评价", "各维度分(json)",
        ]

        data_rows = [row for row in rows[1:] if any(cell.strip() for cell in row)]
        assert any(row[0] == "2024001" and row[1] == "张三" and row[2] == "G1" for row in data_rows)
        evaluated_row = next(
            row for row in data_rows if row[0] == "2024001" and row[1] == "张三"
        )
        assert evaluated_row[3] == "evaluated"
        assert evaluated_row[4] == "B"
        assert evaluated_row[5] == "B"
        assert evaluated_row[6]
        assert "更多功能测试" in evaluated_row[7]
        dims = json.loads(evaluated_row[11])
        assert dims["需求理解"] == 85
        assert dims["实现质量"] == 78
        assert dims["迭代能力"] == 80
        assert evaluated_row[9] == ""


class TestExportWorkbook:
    def test_export_xlsx_is_a_formatted_workbook(self, client, settings):
        _setup_course_and_assignment(client)
        db = SessionLocal()
        assignment = db.query(models.Assignment).one()
        student = db.query(models.Student).order_by(models.Student.student_no).first()
        from app.security import hash_token, new_submit_token
        token = new_submit_token()
        student.submit_token_hash = hash_token(token)
        db.commit()
        db.close()
        assert _upload(client, token, assignment.code, student.student_no).status_code == 201
        db = SessionLocal()
        run_worker_once(db, FakeLLMProvider(responses=[_valid_individual_json()]), settings)
        db.close()

        response = client.get(f"/assignments/{assignment.id}/export.xlsx")
        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        workbook = load_workbook(io.BytesIO(response.content))
        assert workbook.sheetnames == ["反馈总表", "个人反馈", "评分维度", "小组反馈"]
        assert workbook["反馈总表"]["A1"].value.startswith("Vibe 作业反馈表")
        assert workbook["反馈总表"].freeze_panes == "A5"
        assert workbook["个人反馈"]["D5"].alignment.wrap_text is True
